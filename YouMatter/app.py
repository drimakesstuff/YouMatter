from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)

# ------------------- FILE PATH -------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, 'bookings.xlsx')

# ------------------- INITIALIZE EXCEL -------------------
def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Appointments"
        ws.append(["Date", "Time", "Name", "Student ID", "Class Year"])  # Header row
        wb.save(EXCEL_FILE)

initialize_excel()

# ---------------------- ROUTES ----------------------

@app.route('/')
def home():
    return render_template('youmatter.html')

@app.route('/youmatter')
def youmatter():
    return render_template('youmatter.html')

@app.route('/book', methods=['GET'])
def book_page():
    return render_template('book.html')

@app.route('/book', methods=['POST'])
def book():
    data = request.get_json()
    date = data.get('date')
    time = data.get('time')
    name = data.get('name')
    student_id = data.get('student_id')
    class_year = data.get('class_year')

    # Basic validation
    if not all([date, time, name, student_id, class_year]):
        return jsonify({"status": "error", "message": "Missing booking information."}), 400

    # ----------------- DATE VALIDATION -----------------
    try:
        booking_date = datetime.strptime(date, "%Y-%m-%d").date()
        today = datetime.today().date()

        if booking_date < today:
            return jsonify({"status": "error", "message": "You cannot book a past date."}), 400

    except ValueError:
        return jsonify({"status": "error", "message": "Invalid date format."}), 400

    # ----------------- LOAD AND CHECK EXCEL -----------------
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Check if slot already booked
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == date and row[1] == time:
            return jsonify({"status": "error", "message": "This slot is already booked."}), 409

    # Add new booking
    ws.append([date, time, name, student_id, class_year])
    wb.save(EXCEL_FILE)

    return jsonify({"status": "success", "message": "Appointment booked!"})

@app.route('/get_slots', methods=['GET'])
def get_slots():
    if not os.path.exists(EXCEL_FILE):
        return jsonify({})

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    bookings = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        date, time, *_ = row
        if date not in bookings:
            bookings[date] = []
        bookings[date].append(time)

    return jsonify(bookings)

# ------------------ Emotion Pages ------------------

@app.route('/vent')
def vent():
    return render_template('vent.html')

@app.route('/sad')
def sad():
    return render_template('sad.html')

@app.route('/angry')
def angry():
    return render_template('angry.html')

@app.route('/anxious')
def anxious():
    return render_template('anxious.html')

@app.route('/calm')
def calm():
    return render_template('calm.html')

@app.route('/tired')
def tired():
    return render_template('tired.html')

@app.route('/happy')
def happy():
    return render_template('happy.html')

# ----------- After Emotion Pages ------------------

@app.route('/afterangry')
def after_angry():
    return render_template('afterangry.html')

@app.route('/afteranxious')
def after_anxious():
    return render_template('afteranxious.html')

@app.route('/aftersad')
def after_sad():
    return render_template('aftersad.html')

@app.route('/aftercalm')
def after_calm():
    return render_template('aftercalm.html')

@app.route('/aftertired')
def after_tired():
    return render_template('aftertired.html')

@app.route('/afterhappy')
def after_happy():
    return render_template('afterhappy.html')

@app.route('/askforhelp')
def askforhelp():
    return render_template('askforhelp.html')

# ------------------- RUN APP -----------------------
if __name__ == '__main__':
    app.run(debug=True)
